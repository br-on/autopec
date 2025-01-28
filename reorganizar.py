import os
import openpyxl

def reorganizar_colunas(input_dir, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Percorrer todos os arquivos .xlsx na pasta input_dir
    for filename in os.listdir(input_dir):
        # Ignorar arquivos temporários do Excel
        if filename.startswith('~$') or not filename.endswith(".xlsx"):
            continue

        input_path = os.path.join(input_dir, filename)
        output_path = os.path.join(output_dir, filename)

        try:
            # Abrir o arquivo .xlsx
            wb = openpyxl.load_workbook(input_path)
            ws = wb.active

            # Percorrer todas as linhas da planilha
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                # A segunda coluna contém os dados separados por ponto e vírgula
                segunda_coluna = row[1].value
                if segunda_coluna:
                    # Separar os dados da segunda coluna por ponto e vírgula
                    dados_separados = segunda_coluna.split(';')
                    
                    # Preencher as colunas subsequentes com os dados separados
                    for i, valor in enumerate(dados_separados):
                        ws.cell(row=row[0].row, column=3 + i, value=valor)  # Começa da coluna 3

            # Excluir a segunda coluna (contendo os dados separados por vírgula)
            ws.delete_cols(2)

            # Salvar o arquivo com os dados reorganizados
            wb.save(output_path)
            print(f"Arquivo {filename} reorganizado e salvo em {output_path}.")
        except Exception as e:
            print(f"Erro ao processar o arquivo {filename}: {e}")

# Defina os caminhos das pastas de entrada e saída
input_dir = 'tabelas_separadas'  # Pasta onde estão os arquivos .xlsx
output_dir = 'tabelas_reorganizadas'  # Pasta onde os arquivos reorganizados serão salvos

# Chama a função para reorganizar os arquivos
reorganizar_colunas(input_dir, output_dir)
