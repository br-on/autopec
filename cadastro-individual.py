import csv
import os
import chardet
import openpyxl
from openpyxl import Workbook

def detectar_codificacao(arquivo):
    return 'ISO-8859-1'  # Usando uma codificação mais flexível para evitar erros de leitura

def obter_valor_b11(input_csv):
    encoding = detectar_codificacao(input_csv)
    with open(input_csv, 'r', encoding=encoding) as file:
        reader = list(csv.reader(file))
        
        # Verificando quantas linhas e colunas o arquivo tem
        print(f"Linhas no arquivo CSV: {len(reader)}")

        # Inspeção das primeiras 15 linhas do arquivo CSV para depuração
        print("Primeiras linhas do arquivo CSV:")
        for i, row in enumerate(reader[:15]):  # Verificando as 15 primeiras linhas
            print(f"Linha {i+1}: {row}")
        
        # Verifica se o arquivo tem pelo menos 11 linhas
        if len(reader) > 10:
            # A linha B11 está em reader[10], e o valor desejado está após o ';'
            linha_b11 = reader[10][0]  # Pega a string 'Equipe;0000152404 - BERARDO'
            valor_b11 = linha_b11.split(';')[1].strip()  # Divide e pega a segunda parte
        else:
            valor_b11 = "N/A"  # Caso não haja B11, retorna "N/A"
    
    print(f"Valor de B11 lido: '{valor_b11}'")  # Debug
    return valor_b11

def separar_tabelas(input_csv, output_dir, titulos):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Detectando a codificação do arquivo de entrada
    encoding = detectar_codificacao(input_csv)

    # Obter o valor da célula B11
    valor_b11 = obter_valor_b11(input_csv)

    with open(input_csv, 'r', encoding=encoding) as file:
        reader = list(csv.reader(file))
        tabela_atual = []
        titulo_atual = None
        
        for i, row in enumerate(reader):
            if not row:
                continue

            if row[0] in titulos:  # Verifica se a linha contém um título válido
                if tabela_atual and titulo_atual:  # Salva a tabela anterior
                    salvar_tabela(titulo_atual, tabela_atual, output_dir, valor_b11)
                titulo_atual = row[0]  # Atualiza o título atual
                tabela_atual = []  # Reinicia a tabela
            elif row:  # Adiciona linha à tabela atual
                tabela_atual.append(row)

        # Salva a última tabela, se existir
        if tabela_atual and titulo_atual:
            salvar_tabela(titulo_atual, tabela_atual, output_dir, valor_b11)

def salvar_tabela(titulo, tabela, output_dir, valor_b11):
    nome_arquivo = titulo.replace(" / ", "_").replace(" ", "_").replace("-", "_").replace("__", "_").strip("_") + '.xlsx'
    caminho_completo = os.path.join(output_dir, nome_arquivo)

    # Cria uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tabela"

    # Adiciona a coluna B11 à esquerda de todas as linhas da tabela
    for i, row in enumerate(tabela, start=1):
        ws.cell(row=i, column=1, value=valor_b11)  # Coloca o valor de B11 na coluna 1
        for j, cell in enumerate(row, start=2):  # Adiciona os dados da tabela a partir da coluna 2
            ws.cell(row=i, column=j, value=cell)

    # Salva a planilha
    wb.save(caminho_completo)
    print(f"Tabela '{titulo}' salva em '{caminho_completo}'.")

# Configurações do script
input_csv = 'Relatório de cadastro individual-20250128133355.csv'
output_dir = 'tabelas_separadas'

# Títulos das tabelas
titulos = [
    "Identificação do usuário / cidadão - Faixa etária",
    "Identificação do usuário / cidadão",
    "Identificação do usuário / cidadão - Sexo",
    "Identificação do usuário / cidadão - Raça / Cor",
    "Identificação do usuário / cidadão - Etnia",
    "Identificação do usuário / cidadão - Nacionalidade",
    "Informações sociodemográficas - Relação de parentesco com o responsável familiar",
    "Informações sociodemográficas - Ocupação",
    "Informações sociodemográficas - Qual é o curso mais elevado que frequenta ou frequentou",
    "Informações sociodemográficas - Situação no mercado de trabalho",
    "Informações sociodemográficas - Crianças de 0 a 9 anos, com quem fica",
    "Informações sociodemográficas - Orientação sexual",
    "Informações sociodemográficas - Identidade de gênero",
    "Informações sociodemográficas - Deficiência",
    "Informações sociodemográficas - Povos e comunidades",
    "Outras informações sociodemográficas",
    "Condições / Situações de saúde gerais",
    "Condições / Situações de saúde gerais - Sobre seu peso, você se considera",
    "Condições / Situações de saúde gerais - Doença respiratória",
    "Condições / Situações de saúde gerais - Doença cardíaca",
    "Condições / Situações de saúde gerais - Problemas nos rins",
    "Cidadão em situação de rua",
    "Cidadão em situação de rua - Tempo em situação de rua",
    "Cidadão em situação de rua - Quantas vezes se alimenta ao dia",
    "Cidadão em situação de rua - Qual a origem da alimentação",
    "Cidadão em situação de rua - Tem acesso à higiene pessoal",
    "TRIA - Nos últimos três meses, os alimentos acabaram antes que a pessoa tivesse dinheiro para comprar mais comida?",
    "TRIA - Nos últimos três meses, a pessoa comeu apenas alguns alimentos que ainda tinha, porque o dinheiro acabou?"
]

# Passo 1: Separar as tabelas
separar_tabelas(input_csv, output_dir, titulos)
